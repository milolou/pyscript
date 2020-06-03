# blankRowInserter.py - insert a certain number of blanks into a specific row.

import openpyxl
from openpyxl.utils import get_column_letter

def blankInserter(N,M,filename):
    wb = openpyxl.load_workbook(filename)
    nwb = openpyxl.Workbook()
    sheet = wb.active
    nsheet = nwb.active
    for i in range(N):
        for cellObj in list(sheet.rows)[i]:
            nsheet[cellObj.coordinate].value = cellObj.value
    for x in range(N,sheet.max_row):
        for newObj in list(sheet.rows)[x]:
            nsheet[get_column_letter(newObj.column) + '%s'%(x + 1 + M)].value = newObj.value
    nwb.save('inserted.xlsx')
    wb.close()
            
blankInserter(3,2,'example.xlsx')   
'''   
#! python3
import openpyxl
from openpyxl.utils import get_column_letter

def rowCopy (ws1,row1,ws2,row2):    #定义行复制函数，以减少冗余代码
    for cell in ws1[row1]:
        ws2[get_column_letter(cell.column) + str(row2)].value = cell.value
    return 0

def blankRowInserter (N, M, fileName):    #以行为操作单位，而不是直接遍历所有单元格，避免了在for循环中使用过多判断
    wb1 = openpyxl.load_workbook(fileName)
    wb2 = openpyxl.Workbook()
    ws1 = wb1.active
    ws2 = wb2.active
    if N == 1:    # range(1,1)这种特殊情况单独处理
        for rowNum in range(1,ws1.max_row + 1):
            rowCopy(ws1,rowNum,ws2,rowNum + M)
    elif N > 1 and N <= ws1.max_row:
        for rowNum in range(1,N):
            rowCopy(ws1,rowNum,ws2,rowNum)
        for rowNum in range(N,ws1.max_row + 1):
            rowCopy(ws1,rowNum,ws2,rowNum + M)
    else:
        return 0
    wb1.close()
    wb2.save('new' + fileName)

N = int(input('please input the row number: '))
M = int(input('how many lines would you like to insert: '))
fileName = input('please input the excel file name: ')

blankRowInserter(3, 2,'example.xlsx')'''
