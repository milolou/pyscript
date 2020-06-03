# textToSpreadsheet.py - copy the text in textFiles to Spreadsheet.

import openpyxl
from openpyxl.utils import get_column_letter

fileList = ['errorInfo.txt','debug.log','replacedText.txt']

def copytext(textFileList):
    wb = openpyxl.Workbook()
    sheet = wb.active
    startColumn = sheet.max_column
    for textFile in textFileList:
        newFile = open(textFile)
        textList = newFile.readlines()
        for text in textList:
            sheet[get_column_letter(startColumn) + '%s'%(textList.index(text)+1)].value = text
        newFile.close()
        startColumn = sheet.max_column + 1
    wb.save('textSpreadsheet.xlsx')

copytext(fileList)
