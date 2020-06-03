# spreadsheetInverter.py - invert the cell by coordinate.

import openpyxl
from openpyxl.utils import get_column_letter

def inverter(filename,newfilename):
    wb = openpyxl.load_workbook(filename)
    nwb = openpyxl.Workbook()
    sheet = wb.active
    nsheet = nwb.active
    m = sheet.max_column
    n = sheet.max_row
    sheetData = []
    for i in range(1,m+1):
        sheetData.append([])
        for j in range(1,n+1):
            sheetData[i-1].append(sheet.cell(row=j,column=i).value)
    for x in range(1,n+1):
        for y in range(1,m+1):
            nsheet.cell(row=y,column=x).value = sheetData[y-1][x-1]
    wb.close()
    nwb.save(newfilename)

inverter('example_copy.xlsx','inverted.xlsx')
