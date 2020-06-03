# excelToCsvConverter.py - Convert spreadsheets to csv file.

import openpyxl,os,csv

os.makedirs('csvFile',exist_ok=True)
for excelFile in os.listdir('.\\excelSpreadsheets'):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(os.path.join('.\\excelSpreadsheets',excelFile))
    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheetName]

        # Create the CSV filename from the Excel filename and sheet title.
        basename = os.path.basename(excelFile)
        l = len(basename)
        fileName = basename[:(l-5)]
        csvName = fileName + '_' + sheetName + '.csv'
        csvFile = open(os.path.join('.\\csvFile',csvName),'w',newline='')
        
        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFile)
        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum,column=colNum).value)

            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)

        csvFile.close()
