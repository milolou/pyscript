# multiplicationTable.py - make a table for multiplication.

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Define a function to do the work.
def tableMaker(N):
    # Create a workbook.
    wb = openpyxl.Workbook()
    sheet = wb.active
    # Populate the cells for calculating.
    boldFont = Font(bold=True)
    for i in range(N):
        sheet['A' + str(i + 2)].value = i + 1
        sheet['A' + str(i + 2)].font = boldFont
        sheet[get_column_letter(i + 2) + '1'].value = i + 1
        sheet[get_column_letter(i + 2) + '1'].font = boldFont
    # Populate the other cells by calculating.
    topLeftCell = 'B2'
    bottomRightCell = get_column_letter(N + 1) + '%s'%(N + 1)
    leftedCells = tuple(sheet[topLeftCell:bottomRightCell])
    for rowOfCellObjects in leftedCells:
        for cellObj in rowOfCellObjects:
            upper_value = get_column_letter(cellObj.column) + '1'
            left_value = 'A' + str(cellObj.row)
            cellObj.value = '=%s*%s '%(upper_value,left_value)
    # Save the file.
    wb.save('multiplicatioTable.xlsx')

tableMaker(6)
